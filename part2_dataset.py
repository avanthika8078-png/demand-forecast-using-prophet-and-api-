import pandas as pd
import numpy as np
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error
from tabulate import tabulate
import joblib
import kagglehub
path = kagglehub.dataset_download("raminhuseyn/demand-forecasting-dataset")

df=pd.read_csv(f"{path}/demand_forecasting.csv")
df.head()

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA   = f"{path}/demand_forecasting.csv"
OUTPUT = "/mnt/user-data/outputs/weekly_sales_report.xlsx"

#  1. LOAD & PREP
df = pd.read_csv(DATA, parse_dates=["Date"])
df["Revenue"]    = df["Demand"] * df["Price"]
df["WeekStart"]  = df["Date"].dt.to_period("W").apply(lambda p: p.start_time)
df["WeekLabel"]  = df["WeekStart"].dt.strftime("W/C %d %b %Y")

#  2. BUILD TABLES
# a) Weekly summary
weekly = (df.groupby(["WeekLabel", "WeekStart"])
            .agg(
                Transactions = ("Demand",   "count"),
                Units_Sold    = ("Demand",   "sum"),
                Revenue      = ("Revenue",  "sum"),
                Avg_Price    = ("Price",    "mean"),
                Promotions   = ("Promotion","sum"),
            )
            .reset_index()
            .sort_values("WeekStart")
            .drop(columns="WeekStart"))
weekly.columns = ["Week", "Transactions", "Units Sold",
                  "Revenue ($)", "Avg Price ($)", "Promo Rows"]
weekly["Avg Price ($)"] = weekly["Avg Price ($)"].round(2)

# b) Category breakdown
cat = (df.groupby("Category")
         .agg(
             Transactions = ("Demand",    "count"),
             Units_Sold   = ("Demand",    "sum"),
             Revenue      = ("Revenue",   "sum"),
             Avg_Demand   = ("Demand",    "mean"),
             Avg_Price    = ("Price",     "mean"),
             Promo_Pct    = ("Promotion", "mean"),
         )
         .reset_index()
         .sort_values("Revenue", ascending=False))
cat["Avg_Demand"]  = cat["Avg_Demand"].round(1)
cat["Avg_Price"]   = cat["Avg_Price"].round(2)
cat["Promo_Pct"]   = (cat["Promo_Pct"] * 100).round(1)
cat.columns = ["Category", "Transactions", "Units Sold", "Revenue ($)",
               "Avg Demand", "Avg Price ($)", "% Promoted"]

# c) Region breakdown
region = (df.groupby("Region")
            .agg(
                Units_Sold = ("Demand",    "sum"),
                Revenue    = ("Revenue",   "sum"),
                Avg_Price  = ("Price",     "mean"),
                Promo_Pct  = ("Promotion", "mean"),
            )
            .reset_index()
            .sort_values("Revenue", ascending=False))
region["Avg_Price"] = region["Avg_Price"].round(2)
region["Promo_Pct"] = (region["Promo_Pct"] * 100).round(1)
region.columns = ["Region", "Units Sold", "Revenue ($)", "Avg Price ($)", "% Promoted"]

# d) Top products
top_prod = (df.groupby("Product ID")
              .agg(Units_Sold=("Demand", "sum"),
                   Revenue   =("Revenue","sum"))
              .reset_index()
              .sort_values("Revenue", ascending=False)
              .head(10))
top_prod.columns = ["Product ID", "Units Sold", "Revenue ($)"]

#  3. EXCEL STYLES
NAVY  = "1F3864"; BLUE  = "2E75B6"; LTBLUE = "D6E4F0"
WHITE = "FFFFFF"; GREY  = "F2F2F2"

thin   = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center")
right  = Alignment(horizontal="right",  vertical="center")

def hdr(ws, row, col, value):
    c = ws.cell(row, col, value)
    c.font      = Font(bold=True, color=WHITE, name="Arial", size=10)
    c.fill      = PatternFill("solid", fgColor=NAVY)
    c.alignment = center
    c.border    = border
    return c

def sec_title(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font      = Font(bold=True, color=WHITE, name="Arial", size=12)
    c.fill      = PatternFill("solid", fgColor=BLUE)
    c.alignment = center
    ws.row_dimensions[row].height = 22

def write_table(ws, dataframe, start_row):
    ncols = len(dataframe.columns)
    # Header
    for ci, col in enumerate(dataframe.columns, 1):
        hdr(ws, start_row, ci, col)
    # Data rows
    for ri, row_data in enumerate(dataframe.itertuples(index=False), 1):
        fill = PatternFill("solid", fgColor=LTBLUE if ri % 2 == 0 else WHITE)
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(start_row + ri, ci, val)
            c.fill   = fill
            c.border = border
            col_name = dataframe.columns[ci - 1]
            if "Revenue" in col_name:
                c.number_format = '$#,##0.00'
                c.alignment = right
            elif "Price" in col_name:
                c.number_format = '$#,##0.00'
                c.alignment = right
            elif col_name in ("Units Sold", "Transactions", "Units Ordered", "Promotions", "Promo Rows"):
                c.number_format = '#,##0'
                c.alignment = right
            elif "%" in col_name:
                c.number_format = '0.0"%"'
                c.alignment = right
            else:
                c.alignment = left
    return start_row + len(dataframe) + 2

def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 38)

#  4. BUILD WORKBOOK
wb = Workbook()

#  Sheet 1: Weekly Summary
ws1 = wb.active
ws1.title = "Weekly Summary"
ws1.sheet_view.showGridLines = False

# Title banner
ws1.merge_cells("A1:F2")
t = ws1["A1"]
t.value     = " Weekly Sales Report  —  Demand Forecasting Dataset"
t.font      = Font(bold=True, color=WHITE, name="Arial", size=15)
t.fill      = PatternFill("solid", fgColor=NAVY)
t.alignment = center
ws1.row_dimensions[1].height = 28
ws1.row_dimensions[2].height = 18

# KPI row
total_rev  = df["Revenue"].sum()
total_units= df["Demand"].sum()
promo_pct  = df["Promotion"].mean() * 100
avg_price  = df["Price"].mean()
kpis = [
    ("Total Revenue",   f"${total_rev:,.0f}"),
    ("Total Units Sold",f"{total_units:,}"),
    ("Avg Unit Price",  f"${avg_price:.2f}"),
    ("% Rows Promoted", f"{promo_pct:.1f}%"),
]
ws1.append([]); ws1.append([])   # rows 3-4 spacer
for ki, (label, val) in enumerate(kpis, 1):
    lc = ws1.cell(4, ki * 2 - 1, label)
    vc = ws1.cell(4, ki * 2, val)
    lc.font = Font(bold=True, name="Arial", size=10, color=NAVY)
    vc.font = Font(bold=True, name="Arial", size=10, color=BLUE)
ws1.append([])

sec_title(ws1, 6, len(weekly.columns), "Weekly Revenue & Demand Summary")
write_table(ws1, weekly, 7)
auto_width(ws1)

#  Sheet 2: Category Breakdown
ws2 = wb.create_sheet("Category Breakdown")
ws2.sheet_view.showGridLines = False
sec_title(ws2, 1, len(cat.columns), "Sales Performance by Product Category")
write_table(ws2, cat, 2)
auto_width(ws2)

#  Sheet 3: Region Breakdown
ws3 = wb.create_sheet("Region Breakdown")
ws3.sheet_view.showGridLines = False
sec_title(ws3, 1, len(region.columns), "Sales Performance by Region")
write_table(ws3, region, 2)
auto_width(ws3)

#  Sheet 4: Top Products
ws4 = wb.create_sheet("Top Products")
ws4.sheet_view.showGridLines = False
sec_title(ws4, 1, len(top_prod.columns), "Top 10 Products by Revenue")
write_table(ws4, top_prod, 2)
auto_width(ws4)

wb.save(OUTPUT)
print(f"✓ Weekly sales report saved → {OUTPUT}")
